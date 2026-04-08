import os
import time
import argparse
import logging
import hashlib
import re
import uuid
import threading
import datetime
import atexit
import json
from logging.handlers import RotatingFileHandler
from typing import Dict, List, Tuple, Optional, Final, Any, Union, IO, BinaryIO
import multiprocessing
import base64
from cryptography.fernet import Fernet
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
import questionary
import pypdf
import docx
import openpyxl
import extract_msg
import zipfile
import tarfile
from email import policy
from email.parser import BytesParser
import pandas as pd
import io

from dataclasses import dataclass

@dataclass(frozen=True)
class Config:
    """Centralized configuration for TraceWords"""
    MAX_FILE_SIZE_MB: Final[int] = 100
    MAX_TOTAL_EXTRACT_SIZE_MB: Final[int] = 500
    MAX_ARCHIVE_DEPTH: Final[int] = 3
    MAX_ARCHIVE_FILES: Final[int] = 1000
    CHUNK_SIZE: Final[int] = 8192
    REGEX_TIMEOUT_SEC: Final[int] = 2
    LOG_FILE: Final[str] = "tracewords_info.log"
    AUDIT_LOG_FILE: Final[str] = "tracewords_audit.log"
    PRIVACY_LOG_FILE: Final[str] = "tracewords_privacy.log"
    SUPPORTED_FORMATS: Final[Tuple[str, ...]] = (
        ".txt", ".json", ".csv", ".log", ".xml", ".html", ".py", ".js", 
        ".php", ".sql", ".conf", ".ini", ".pdf", ".docx", ".xlsx", ".zip", 
        ".tar.gz", ".eml", ".msg"
    )
    HASH_ALGORITHM: Final[str] = "sha256"
    FILE_SCAN_TIMEOUT_SEC: Final[int] = 60
    SECURE_DELETE_PASSES: Final[int] = 3
    ZIP_BOMB_RATIO_LIMIT: Final[int] = 100
    LARGE_FILE_THRESHOLD_BYTES: Final[int] = 10 * 1024 * 1024  # 10MB
    LOG_MAX_BYTES: Final[int] = 50 * 1024 * 1024  # 50MB
    LOG_BACKUP_COUNT: Final[int] = 5
    
    @property
    def max_file_size_bytes(self) -> int:
        return self.MAX_FILE_SIZE_MB * 1024 * 1024
    
    @property
    def max_total_extract_size_bytes(self) -> int:
        return self.MAX_TOTAL_EXTRACT_SIZE_MB * 1024 * 1024

CONFIG = Config()
CONSOLE = Console()

# Global Thread Pool for shared tasks
_EXECUTOR = ThreadPoolExecutor(max_workers=min(32, (multiprocessing.cpu_count() or 1) * 4))

def cleanup_executor():
    """Cleanup global executor on program exit"""
    _EXECUTOR.shutdown(wait=True)
    logging.info("Global ThreadPoolExecutor cleaned up")

atexit.register(cleanup_executor)

class ArchiveStats:
    """Thread-safe archive statistics tracking"""
    def __init__(self):
        self._total_size = 0
        self._total_files = 0
        self._lock = threading.Lock()
    
    def add_file(self, size: int) -> bool:
        with self._lock:
            if (self._total_size + size) > CONFIG.max_total_extract_size_bytes:
                return False
            self._total_files += 1
            self._total_size += size
            return True
            
    def reset(self):
        with self._lock:
            self._total_size = 0
            self._total_files = 0
            
    @property
    def total_size(self):
        with self._lock:
            return self._total_size
            
    @property
    def total_files(self):
        with self._lock:
            return self._total_files

# Global state for archive tracking
archive_stats = ArchiveStats()

# GDPR/CCPA Compliance Settings
PRIVACY_SETTINGS = {
    "enable_pii_masking": True,
    "enable_data_minimization": True,
    "enable_audit_logging": True,
    "retention_period_days": 90,
    "anonymize_results": False,
    "require_consent": True,
    "enable_right_to_be_forgotten": True
}

# PII Detection Regex Patterns
PII_PATTERNS = {
    'email': r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b',
    'phone_us': r'\b(?:\+?1[-.\s]?)?\(?([0-9]{3})\)?[-.\s]?([0-9]{3})[-.\s]?([0-9]{4})\b',
    'phone_international': r'\+[1-9]\d{1,14}',
    'ssn_us': r'\b(?!000|666|9\d{2})\d{3}[-.]?(?!00)\d{2}[-.]?(?!0000)\d{4}\b',
    'credit_card': r'\b(?:4[0-9]{12}(?:[0-9]{3})?|5[1-5][0-9]{14}|6(?:011|5[0-9]{2})[0-9]{12}|3[47][0-9]{13}|3[0-9]{13}|2[0-9]{15})\b',
    # v5.0 Improved IP regex (preventing 999.999...)
    'ip_address': r'\b(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\b',
    'mac_address': r'\b([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})\b',
    'url': r'https?://(?:[-\w.])+(?:[:\d]+)?(?:/(?:[\w/_.])*(?:\?(?:[\w&=%.])*)?(?:#(?:[\w.])*)?)?',
    'tc_kimlik': r'(?<![\d])([1-9][0-9]{10})(?![\d])', 
    'iban': r'\b[A-Z]{2}[0-9]{2}[A-Z0-9]{4}[0-9]{7}([A-Z0-9]?){0,16}\b',
    'passport': r'\b[A-Z][0-9]{8}\b',
    'date_of_birth': r'\b(0[1-9]|[12][0-9]|3[01])[\/\-\.](0[1-9]|1[012])[\/\-\.]((19|20)\d\d)\b'
}

# Utility Functions
def sanitize_path(path: str) -> str:
    """Secure path normalization and traversal protection"""
    if not path:
        return "unknown"
    try:
        # Resolve to absolute path and normalize
        abs_path = os.path.abspath(path)
        cwd = os.path.abspath('.')
        
        # Check if it's within CWD, if not, just take the basename
        if not (abs_path == cwd or abs_path.startswith(cwd + os.sep)):
            return os.path.basename(path)
        return abs_path
    except Exception:
        return os.path.basename(path)

def validate_tc_kimlik(tc: str) -> bool:
    """TC Kimlik No verification algorithm"""
    if not tc or len(tc) != 11 or not tc.isdigit() or tc[0] == '0':
        return False
    
    digits = [int(d) for d in tc]
    
    # 10. digit control
    sum_odd = sum(digits[0:9:2])  # 1,3,5,7,9
    sum_even = sum(digits[1:8:2])  # 2,4,6,8
    check_10 = (sum_odd * 7 - sum_even) % 10
    if check_10 != digits[9]:
        return False
    
    # 11. digit control
    check_11 = sum(digits[:10]) % 10
    return check_11 == digits[10]

def safe_regex_search(pattern: Any, text: str, timeout_sec: Optional[int] = None) -> List:
    """ReDoS protected regex search using global thread pool"""
    if timeout_sec is None:
        timeout_sec = CONFIG.REGEX_TIMEOUT_SEC
        
    def _search():
        if isinstance(pattern, re.Pattern):
            return pattern.findall(text)
        return re.findall(pattern, text, re.IGNORECASE)

    future = _EXECUTOR.submit(_search)
    try:
        return future.result(timeout=timeout_sec)
    except TimeoutError:
        logging.warning(f"Regex timeout: pattern might be too complex")
        return []
    except Exception as e:
        logging.error(f"Regex error: {e}")
        return []

def secure_delete(filepath: str, passes: int = CONFIG.SECURE_DELETE_PASSES):
    """Secure file deletion (Overwrite with random data multiple times before delete)"""
    if not os.path.exists(filepath):
        return
    try:
        file_size = os.path.getsize(filepath)
        if file_size == 0:
            os.remove(filepath)
            return

        with open(filepath, 'r+b', buffering=0) as f:
            for _ in range(passes):
                f.seek(0)
                # Use chunks for writing random data to avoid memory issues with huge files
                remaining = file_size
                while remaining > 0:
                    write_size = min(remaining, 1024 * 1024) # 1MB chunks
                    f.write(os.urandom(write_size))
                    remaining -= write_size
                f.flush()
                os.fsync(f.fileno())
        os.remove(filepath)
    except Exception as e:
        logging.error(f"Secure delete failed for {filepath}: {e}")
        try:
            os.remove(filepath)
        except OSError as remove_err:
            logging.error(f"Fallback remove also failed for {filepath}: {remove_err}")

class GDPRCompliantStorage:
    """GDPR compliant data storage with automated machine-linked encryption"""
    def _generate_and_save_key(self, key_file: str) -> bytes:
        key = Fernet.generate_key()
        os.makedirs(os.path.dirname(key_file), exist_ok=True)
        with open(key_file, 'wb') as f:
            f.write(key)
        try:
            os.chmod(key_file, 0o600)
        except Exception as e:
            logging.warning(f"Could not set custom permissions on key file: {e}")
        return key

    def __init__(self, key: Optional[bytes] = None):
        if key:
            self.key = key
        else:
            key_env = os.environ.get('TRACEWORDS_ENCRYPTION_KEY')
            key_file = os.path.expanduser('~/.tracewords/keyfile')
            
            if key_env:
                try:
                    candidate = key_env.encode()
                    Fernet(candidate)   # doğrulama için anlık deneme
                    self.key = candidate
                except (ValueError, Exception) as e:
                    logging.error(f"TRACEWORDS_ENCRYPTION_KEY geçersiz Fernet key: {e}. Yeni key üretiliyor.")
                    self.key = self._generate_and_save_key(key_file)
            elif os.path.exists(key_file):
                try:
                    with open(key_file, 'rb') as f:
                        candidate = f.read()
                    Fernet(candidate)   # doğrulama
                    self.key = candidate
                except Exception as e:
                    logging.error(f"Keyfile geçersiz veya bozuk: {e}. Yeni key üretiliyor.")
                    self.key = self._generate_and_save_key(key_file)
            else:
                self.key = self._generate_and_save_key(key_file)
        self.cipher = Fernet(self.key)
    
    def store_encrypted(self, data: str, filepath: str) -> bool:
        """Store data encrypted"""
        try:
            encrypted = self.cipher.encrypt(data.encode())
            with open(filepath, 'wb') as f:
                f.write(encrypted)
            return True
        except Exception as e:
            logging.error(f"Encryption failed: {e}")
            return False

    def store_encrypted_bytes(self, data: bytes, filepath: str) -> bool:
        """Store raw bytes encrypted"""
        try:
            encrypted = self.cipher.encrypt(data)
            with open(filepath, 'wb') as f:
                f.write(encrypted)
            return True
        except Exception as e:
            logging.error(f"Byte encryption failed: {e}")
            return False

    def decrypt_data(self, encrypted_data: bytes) -> str:
        """Decrypt data"""
        try:
            return self.cipher.decrypt(encrypted_data).decode()
        except Exception as e:
            logging.error(f"Decryption failed: {e}")
            return ""

# Logging configuration
_LOGGING_LOCK = threading.Lock()
_LOGGING_INITIALIZED = False

def setup_logging() -> Tuple[logging.Logger, logging.Logger]:
    """Setup multiple log systems (standard, audit, privacy) - v5.0: Robust initialization"""
    global _LOGGING_INITIALIZED
    
    with _LOGGING_LOCK:
        if _LOGGING_INITIALIZED:
            return logging.getLogger('audit'), logging.getLogger('privacy')
            
        # Root logger
        logger = logging.getLogger()
        audit_logger = logging.getLogger('audit')
        privacy_logger = logging.getLogger('privacy')

        # Ensure root logger has a RotatingFileHandler for tracewords_info.log.
        # NOTE: We check handler types explicitly instead of hasHandlers() because
        # Python may have already attached a StreamHandler (via logging.lastResort or
        # an early logging.info() call in __main__) before setup_logging() is invoked,
        # which would cause hasHandlers() to return True and skip file handler setup,
        # leaving tracewords_info.log never created.
        has_file_handler = any(
            isinstance(h, logging.FileHandler) for h in logger.handlers
        )
        if not has_file_handler:
            root_handler = RotatingFileHandler(
                CONFIG.LOG_FILE, 
                maxBytes=CONFIG.LOG_MAX_BYTES, 
                backupCount=CONFIG.LOG_BACKUP_COUNT, 
                encoding="utf-8"
            )
            root_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
            root_handler.setFormatter(root_formatter)
            logger.setLevel(logging.INFO)
            logger.addHandler(root_handler)
        
        # Audit log setup
        audit_logger.setLevel(logging.INFO)
        audit_logger.propagate = False # REC-H: Prevent propagation to root logger
        if not audit_logger.handlers:
            audit_handler = RotatingFileHandler(
                CONFIG.AUDIT_LOG_FILE, 
                maxBytes=CONFIG.LOG_MAX_BYTES, 
                backupCount=CONFIG.LOG_BACKUP_COUNT, 
                encoding='utf-8'
            )
            audit_formatter = logging.Formatter('%(asctime)s - [AUDIT] - %(message)s')
            audit_handler.setFormatter(audit_formatter)
            audit_logger.addHandler(audit_handler)
        
        # Privacy log setup
        privacy_logger.setLevel(logging.INFO)
        privacy_logger.propagate = False # REC-H: Prevent propagation to root logger
        if not privacy_logger.handlers:
            privacy_handler = RotatingFileHandler(
                CONFIG.PRIVACY_LOG_FILE, 
                maxBytes=CONFIG.LOG_MAX_BYTES, 
                backupCount=CONFIG.LOG_BACKUP_COUNT, 
                encoding='utf-8'
            )
            privacy_formatter = logging.Formatter('%(asctime)s - [PRIVACY]- %(message)s')
            privacy_handler.setFormatter(privacy_formatter)
            privacy_logger.addHandler(privacy_handler)
        
        _LOGGING_INITIALIZED = True
        return audit_logger, privacy_logger

def generate_session_id() -> str:
    """Generate a unique session ID for each analysis"""
    return str(uuid.uuid4())

def validate_keywords(keywords: List[str]) -> Tuple[bool, str]:
    """Validate keyword inputs"""
    if not keywords:
        return False, "Keyword listesi boş"
    
    if len(keywords) > 100:
        return False, "Maksimum 100 keyword desteklenir"
    
    for kw in keywords:
        if len(kw) < 2:
            return False, f"Keyword çok kısa: '{kw}' (min 2 karakter)"
        if len(kw) > 200:
            return False, f"Keyword çok uzun: '{kw}' (max 200 karakter)"
        
        # Check for injection attempts
        dangerous_chars = [';', '|', '&', '$', '`', '\n', '\r']
        if any(char in kw for char in dangerous_chars):
            return False, f"Geçersiz karakter bulundu: '{kw}'"
    
    return True, ""

def calculate_file_hash(filepath: str, algorithm: str = CONFIG.HASH_ALGORITHM) -> str:
    """Calculate cryptographic hash of a file for forensic integrity verification.

    Args:
        filepath: Absolute or relative path to the target file.
        algorithm: Hash algorithm name (default: sha256). Accepts any
                   hashlib-supported algorithm (e.g. 'sha256', 'sha3_256').

    Returns:
        Hex digest string, or one of: FILE_NOT_FOUND, PERMISSION_DENIED,
        IO_ERROR, UNKNOWN_ERROR on failure.
    """
    try:
        if not os.path.exists(filepath):
            return "FILE_NOT_FOUND"
        
        h = hashlib.new(algorithm)
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(CONFIG.CHUNK_SIZE), b""):
                h.update(chunk)
        return h.hexdigest()
    
    except PermissionError:
        logging.error(f"Erişim reddedildi: {filepath}")
        return "PERMISSION_DENIED"
    except IOError as e:
        logging.error(f"I/O hatası ({filepath}): {e}")
        return "IO_ERROR"
    except Exception as e:
        logging.error(f"Beklenmeyen hash hatası ({filepath}): {e}")
        return "UNKNOWN_ERROR"

def log_privacy_event(privacy_logger: logging.Logger, event_type: str, details: str, session_id: str) -> None:
    """Log privacy events - v5.0: Masked details"""
    # Anonymize details to prevent sensitive info in log
    safe_details = details[:100] + "..." if len(details) > 100 else details
    privacy_logger.info(f"Session: {session_id} | Event: {event_type} | Details: {safe_details}")

def log_audit_event(audit_logger: logging.Logger, action: str, resource: str, user: str, session_id: str, result: str = "SUCCESS") -> None:
    """Log audit events - v5.0: Sanitized resource paths"""
    # Path Traversal Protection for logs
    safe_resource = sanitize_path(resource)
    audit_logger.info(f"Session: {session_id} | Action: {action} | Resource: {safe_resource} | User: {user} | Result: {result}")

def check_consent(session_id: str, audit_logger, batch_mode: bool = False) -> bool:
    """Check user consent for GDPR compliance - v5.0: Batch mode support"""
    if not PRIVACY_SETTINGS["require_consent"] or batch_mode:
        if batch_mode:
            log_audit_event(audit_logger, "CONSENT_AUTO_GRANTED", "batch_mode", "system", session_id)
        return True
    
    CONSOLE.print(Panel.fit(
        "[bold red]GDPR/CCPA UYUMLULUK - VERİ İŞLEME ONAYI[/bold red]\n\n"
        "Bu araç, dosya içeriklerini analiz eder ve kişisel verileri işleyebilir.\n"
        "Aşağıdaki haklar size tanınmıştır:\n"
        "• Verilerinizin işlenmesini reddetme hakkı\n"
        "• Kişisel verilerinizin silinmesini talep etme hakkı\n"
        "• Veri işleme süreçleri hakkında bilgi alma hakkı\n"
        "• Analiz sonuçlarına erişim ve düzeltme hakkı\n\n"
        "Veri işleme süreci:\n"
        "1. Dosyalar analiz edilir ve anahtar kelimeler aranır\n"
        "2. Kişisel veriler otomatik olarak maskelenir (PII)\n"
        "3. Analiz sonuçları belirtilen süre boyunca saklanır\n"
        "4. Audit logları uyumluluk için kaydedilir",
        title="Güvenlik Bilgilendirmesi", border_style="yellow"
    ))
    
    consent = questionary.confirm("Bu koşulları kabul ediyor musunuz?").ask()
    
    if consent:
        log_audit_event(audit_logger, "CONSENT_GRANTED", "data_processing", "user", session_id)
        CONSOLE.print("[green]✅ Veri işleme onayı verildi.[/green]")
        return True
    else:
        log_audit_event(audit_logger, "CONSENT_DENIED", "data_processing", "user", session_id)
        CONSOLE.print("[red]❌ Veri işleme onayı reddedildi. Analiz sonlandırılıyor.[/red]")
        return False

def detect_pii(content: str) -> Dict[str, List[str]]:
    """Detect PII with ReDoS protection and TC Kimlik validation"""
    detected_pii: Dict[str, List[str]] = {}
    
    for pii_type, pattern in PII_PATTERNS.items():
        # Use ReDoS protected search
        matches = safe_regex_search(pattern, content)
        
        if matches:
            if isinstance(matches[0], tuple):  # Group capture case
                matches = [''.join(match) for match in matches]
            
            # Special validation for TC Kimlik
            if pii_type == 'tc_kimlik':
                # Filter matches using the validation algorithm
                matches = [tc for tc in matches if validate_tc_kimlik(tc)]
                if not matches: 
                    continue
                
            detected_pii[pii_type] = list(set(matches))  # Unique values
    
    return detected_pii

def mask_pii(content: str, pii_data: Dict[str, List[str]], session_id: str, privacy_logger) -> str:
    """Mask PII data"""
    if not PRIVACY_SETTINGS["enable_pii_masking"]:
        return content
    
    masked_content = content
    total_masked = 0
    
    for pii_type, pii_values in pii_data.items():
        for pii_value in pii_values:
            if pii_type == 'email':
                mask = f"[EMAIL_{total_masked}]"
            elif pii_type in ['phone_us', 'phone_international']:
                mask = f"[PHONE_{total_masked}]"
            elif pii_type == 'ssn_us':
                mask = f"[SSN_{total_masked}]"
            elif pii_type == 'credit_card':
                mask = f"[CREDIT_CARD_{total_masked}]"
            elif pii_type == 'ip_address':
                mask = f"[IP_ADDR_{total_masked}]"
            elif pii_type == 'tc_kimlik':
                mask = f"[TC_KIMLIK_{total_masked}]"
            elif pii_type == 'iban':
                mask = f"[IBAN_{total_masked}]"
            else:
                mask = f"[PII_{pii_type.upper()}_{total_masked}]"
            
            masked_content = masked_content.replace(pii_value, mask)
            total_masked += 1
    
    if total_masked > 0:
        log_privacy_event(privacy_logger, "PII_MASKED", f"Masked {total_masked} PII items", session_id)
    
    return masked_content

def anonymize_file_path(filepath: str) -> str:
    """Anonymize file paths"""
    if not PRIVACY_SETTINGS["anonymize_results"]:
        return filepath
    
    filename = os.path.basename(filepath)
    extension = os.path.splitext(filename)[1]
    anonymous_name = f"file_{hashlib.sha256(filepath.encode()).hexdigest()[:8]}{extension}"
    return anonymous_name


def read_file_content(filepath: str, file_stream: Optional[Union[IO, BinaryIO]] = None, 
                      file_ext: Optional[str] = None) -> str:
    """
    Read file content and convert to lowercase.
    file_stream: File-like object (ZIP/TAR stream) - v5.0: Streaming support.
    """
    ext = file_ext or os.path.splitext(filepath)[1].lower()
    try:
        # Check size if it's a physical file
        if not file_stream and os.path.exists(filepath):
            file_size = os.path.getsize(filepath)
            if file_size > CONFIG.max_file_size_bytes:
                logging.warning(f"{os.path.basename(filepath)} too large ({file_size} bytes), skipping.")
                return ""
        
        # Reading strategy based on file extension
        if ext in (".txt", ".log", ".py", ".js", ".php", ".html", ".xml", ".sql", ".conf", ".ini"):
            if file_stream:
                raw = file_stream.read()
                try:
                    return raw.decode("utf-8").lower()
                except UnicodeDecodeError as e:
                    logging.error(f"Geçersiz byte(lar) stream içinde ({filepath}): {e}")
                    raise
            try:
                with open(filepath, "r", encoding="utf-8") as file:
                    return file.read().lower()
            except UnicodeDecodeError as e:
                logging.error(f"Geçersiz byte(lar) dosyada ({filepath}): {e}")
                raise
        elif ext == ".json":
            if file_stream:
                raw = file_stream.read()
                try:
                    data = json.loads(raw.decode("utf-8"))
                except UnicodeDecodeError as e:
                    logging.error(f"Geçersiz byte(lar) JSON stream içinde ({filepath}): {e}")
                    raise
            else:
                try:
                    with open(filepath, "r", encoding="utf-8") as file:
                        data = json.load(file)
                except UnicodeDecodeError as e:
                    logging.error(f"Geçersiz byte(lar) JSON dosyasında ({filepath}): {e}")
                    raise
            return json.dumps(data, indent=2).lower()
        elif ext == ".csv":
            try:
                if file_stream:
                    # v5.0: Encoding resilience for CSV
                    try:
                        df = pd.read_csv(file_stream, encoding="utf-8")
                    except UnicodeDecodeError:
                        file_stream.seek(0)
                        df = pd.read_csv(file_stream, encoding="latin-1")
                else:
                    try:
                        df = pd.read_csv(filepath, encoding="utf-8")
                    except UnicodeDecodeError:
                        df = pd.read_csv(filepath, encoding="latin-1")
                return df.to_string().lower()
            except Exception as e:
                logging.error(f"CSV okunamadı {filepath}: {e}")
                return ""
        elif ext == ".pdf":
            content = ""
            f = file_stream if file_stream else open(filepath, "rb")
            try:
                reader = pypdf.PdfReader(f)
                for page in reader.pages:
                    content += (page.extract_text() or "") + "\n"
            finally:
                if not file_stream: f.close()
            return content.lower()
        elif ext == ".docx":
            f = file_stream if file_stream else filepath
            doc = docx.Document(f)
            content = "\n".join([para.text for para in doc.paragraphs])
            return content.lower()
        elif ext == ".xlsx":
            f = file_stream if file_stream else filepath
            wb = openpyxl.load_workbook(f, data_only=True)
            content = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    content += " ".join([str(cell) for cell in row if cell is not None]) + "\n"
            return content.lower()
        elif ext == ".msg":
            f = file_stream if file_stream else filepath
            # extract_msg might need a seekable stream or bytes
            if file_stream and not hasattr(file_stream, 'seek'):
                f = io.BytesIO(file_stream.read())
            msg = extract_msg.openMsg(f)
            content = f"Subject: {msg.subject}\nFrom: {msg.sender}\nTo: {msg.to}\nDate: {msg.date}\n\n{msg.body}"
            return content.lower()
        elif ext == ".eml":
            if file_stream:
                msg = BytesParser(policy=policy.default).parsebytes(file_stream.read())
            else:
                with open(filepath, 'rb') as f:
                    msg = BytesParser(policy=policy.default).parse(f)
            body = msg.get_body(preferencelist=('plain'))
            content = f"Subject: {msg['subject']}\nFrom: {msg['from']}\nTo: {msg['to']}\nDate: {msg['date']}\n\n{body.get_content() if body else ''}"
            return content.lower()
    except UnicodeDecodeError:
        raise
    except Exception as e:
        logging.error(f"{os.path.basename(filepath)} okunamadı: {e}")
    return ""

def extract_context(content: str, pattern: re.Pattern, context_lines: int = 2, 
                    apply_masking: bool = True, session_id: str = "", 
                    privacy_logger: Optional[logging.Logger] = None, 
                    highlight: bool = False) -> List[Dict[str, Any]]:
    """
    Extract context around keywords and mask PII.
    v5.0: Unified using pre-compiled pattern and efficient search.
    """
    lines = content.split('\n')
    contexts = []
    
    for i, line in enumerate(lines):
        if pattern.search(line):
            start = max(0, i - context_lines)
            end = min(len(lines), i + context_lines + 1)
            context_block = '\n'.join(lines[start:end])
            
            if apply_masking and privacy_logger:
                pii_detected = detect_pii(context_block)
                if pii_detected:
                    context_block = mask_pii(context_block, pii_detected, session_id, privacy_logger)
                    line_pii = detect_pii(line)
                    line = mask_pii(line, line_pii, session_id, privacy_logger)
            
            matched_line = line.strip()
            if highlight:
                try:
                    matched_line = pattern.sub(r"[bold red]\g<0>[/bold red]", matched_line)
                    context_block = pattern.sub(r"[bold red]\g<0>[/bold red]", context_block)
                except re.error:
                    pass  # keep unhighlighted in case of regex error

            contexts.append({
                'line_number': i + 1,
                'context': context_block,
                'matched_line': matched_line
            })
    return contexts

def process_large_file_chunked(filepath: str, keywords: List[str], 
                                   regex_mode: bool = False, exact_match: bool = False,
                                   session_id: str = "", privacy_logger: Optional[logging.Logger] = None) -> Tuple[Dict, List]:
    """Simplified memory-efficient chunked processing for large text files"""
    matches = {}
    all_contexts = []
    
    # Compile patterns once
    patterns = []
    for kw in keywords:
        try:
            if regex_mode:
                p = re.compile(kw, re.IGNORECASE)
            elif exact_match:
                p = re.compile(r'\b' + re.escape(kw) + r'\b', re.IGNORECASE)
            else:
                p = re.compile(re.escape(kw), re.IGNORECASE)
            patterns.append((kw, p))
        except re.error:
            continue
    
    # Read file line by line for memory efficiency with lookahead support
    try:
        context_radius = 2
        buffer_size = context_radius * 2 + 1   # 5
        lines_buffer = [''] * context_radius 
        last_centered_no = 0            # Last processed line number
        i = -1
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                for i, line in enumerate(f):
                    lines_buffer.append(line)
                    
                    # Process only when buffer is full
                    if len(lines_buffer) == buffer_size:
                        center_no = i - context_radius + 1
                        center_line = lines_buffer[context_radius]
                        last_centered_no = center_no
                        
                        for kw_str, pattern in patterns:
                            if safe_regex_search(pattern, center_line):
                                matches[kw_str] = matches.get(kw_str, 0) + 1
                                context_block = "".join(lines_buffer)
                                
                                if privacy_logger and PRIVACY_SETTINGS["enable_pii_masking"]:
                                    pii_detected = detect_pii(context_block)
                                    if pii_detected:
                                        context_block = mask_pii(context_block, pii_detected, session_id, privacy_logger)
                                
                                try:
                                    highlighted_line = pattern.sub(r"[bold red]\g<0>[/bold red]", center_line.strip())
                                    highlighted_context = pattern.sub(r"[bold red]\g<0>[/bold red]", context_block)
                                except re.error:
                                    highlighted_line = center_line.strip()
                                    highlighted_context = context_block
                                
                                all_contexts.append({
                                    'line_number': center_no,
                                    'context': highlighted_context,
                                    'matched_line': highlighted_line
                                })
                    
                    lines_buffer.pop(0)
        except UnicodeDecodeError as e:
            logging.error(f"Geçersiz byte(lar) büyük dosyada ({filepath}): {e}")
            raise

        # Trailing phase: process lines remaining in the buffer after the last centered line
        # To preserve context, we don't pop until we check the matches
        start_no = (i + 1) - len(lines_buffer) + 1
        trailing_idx = (last_centered_no + 1) - start_no if last_centered_no > 0 else 0
        
        while trailing_idx < len(lines_buffer):
            trailing_line = lines_buffer[trailing_idx]
            trailing_line_no = start_no + trailing_idx
            context_block = "".join(lines_buffer)
            
            for kw_str, pattern in patterns:
                if safe_regex_search(pattern, trailing_line):
                    matches[kw_str] = matches.get(kw_str, 0) + 1
                    
                    masked_context = context_block
                    if privacy_logger and PRIVACY_SETTINGS["enable_pii_masking"]:
                        pii_detected = detect_pii(context_block)
                        if pii_detected:
                            masked_context = mask_pii(context_block, pii_detected, session_id, privacy_logger)
                    
                    try:
                        highlighted_line = pattern.sub(r"[bold red]\g<0>[/bold red]", trailing_line.strip())
                        highlighted_context = pattern.sub(r"[bold red]\g<0>[/bold red]", masked_context)
                    except re.error:
                        highlighted_line = trailing_line.strip()
                        highlighted_context = masked_context
                    
                    all_contexts.append({
                        'line_number': trailing_line_no,
                        'context': highlighted_context,
                        'matched_line': highlighted_line
                    })
            
            trailing_idx += 1

    except UnicodeDecodeError:
        raise
    except Exception as e:
        logging.error(f"Failed to read {filepath} chunked: {e}")
    
    return matches, all_contexts

def process_file(
    filepath: str, 
    keywords: List[str], 
    exact_match: bool, 
    regex_mode: bool = False, 
    session_id: str = "", 
    privacy_logger: Optional[logging.Logger] = None, 
    file_obj: Optional[BinaryIO] = None, 
    original_name: Optional[str] = None, 
    depth: int = 0,
    member_size: int = 0
) -> List:
    """
    Scan a single file and return digital evidence.
    v5.0: Unified regex pipeline and archive streaming support.
    """
    display_name = original_name if original_name else os.path.basename(filepath)
    
    # DÜZELTME: çift uzantıyı (.tar.gz) os.path.splitext'ten önce yakala (BUG-13-A)
    if display_name.lower().endswith('.tar.gz'):
        ext = '.tar.gz'
    else:
        ext = os.path.splitext(display_name)[1].lower()
    
    # Archive security check (v5.0)
    if depth > CONFIG.MAX_ARCHIVE_DEPTH:
        logging.warning(f"Max archive depth reached: {display_name}")
        return []

    # Archive check
    if file_obj is None: # This means it's a physical file path
        if ext == ".zip":
            results = []
            try:
                # Zip Bomb check: Check compression ratio (REC-G)
                compression_ratio_limit = CONFIG.ZIP_BOMB_RATIO_LIMIT
                with zipfile.ZipFile(filepath, 'r') as zf:
                    compressed_size = os.path.getsize(filepath)
                    uncompressed_size = sum(info.file_size for info in zf.infolist())
                    ratio = uncompressed_size / compressed_size if compressed_size > 0 else 0
                    
                    if ratio > compression_ratio_limit:
                        logging.warning(f"Zip bomb detected (Ratio: {ratio:.1f}:1): {display_name}")
                        return []
                        
                    if uncompressed_size > CONFIG.max_total_extract_size_bytes:
                        logging.error(f"Archive too large: {uncompressed_size} bytes")
                        return []

                    for member in zf.namelist():
                        if archive_stats.total_files >= CONFIG.MAX_ARCHIVE_FILES: break
                        
                        if any(member.lower().endswith(s_ext) for s_ext in CONFIG.SUPPORTED_FORMATS):
                            with zf.open(member) as f:
                                info = zf.getinfo(member)
                                if not archive_stats.add_file(info.file_size):
                                    logging.error("Max total extraction size reached!")
                                    break
                                
                                m_time = datetime.datetime(*info.date_time).strftime('%Y-%m-%d %H:%M:%S')
                                member_results = process_file(filepath, keywords, exact_match, regex_mode, session_id, privacy_logger, 
                                                               file_obj=f, original_name=member, depth=depth+1, member_size=info.file_size)
                                for res in member_results:
                                    res[3]['modified'] = f"{m_time} (ZIP)"
                                results.extend(member_results)
                return results
            except Exception as e:
                logging.error(f"ZIP okunamadı {display_name}: {e}")
                return []
        elif ext == ".tar.gz":
            results = []
            try:
                # TAR bomb protection
                with tarfile.open(filepath, 'r:gz') as tf:
                    compressed_size = os.path.getsize(filepath)
                    uncompressed_size = sum(m.size for m in tf.getmembers() if m.isfile())
                    ratio = uncompressed_size / compressed_size if compressed_size > 0 else 0
                    
                    if ratio > CONFIG.ZIP_BOMB_RATIO_LIMIT:  # Same limit as ZIP
                        logging.warning(f"Tar bomb detected (Ratio: {ratio:.1f}:1): {display_name}")
                        return []
                    
                    if uncompressed_size > CONFIG.max_total_extract_size_bytes:
                        logging.error(f"Archive too large: {uncompressed_size} bytes")
                        return []

                    for member in tf.getmembers():
                        if archive_stats.total_files >= CONFIG.MAX_ARCHIVE_FILES: break
                        
                        if member.isfile() and any(member.name.lower().endswith(s_ext) for s_ext in CONFIG.SUPPORTED_FORMATS):
                            f = tf.extractfile(member)
                            if f:
                                if not archive_stats.add_file(member.size):
                                    logging.error("Max total extraction size reached!")
                                    break
                                
                                m_time = datetime.datetime.fromtimestamp(member.mtime).strftime('%Y-%m-%d %H:%M:%S')
                                member_results = process_file(filepath, keywords, exact_match, regex_mode, session_id, privacy_logger, 
                                                               file_obj=f, original_name=member.name, depth=depth+1, member_size=member.size)
                                for res in member_results:
                                    res[3]['modified'] = f"{m_time} (TAR)"
                                results.extend(member_results)
                return results
            except Exception as e:
                logging.error(f"TAR.GZ okunamadı {display_name}: {e}")
                return []

    # Final logic for single files (or archive members)
    # v5.0: Check if we should use chunked processing for large text files
    if file_obj is None and ext in (".txt", ".log", ".py", ".js", ".php", ".html", ".xml", ".sql", ".conf", ".ini"):
        file_size = os.path.getsize(filepath) if os.path.exists(filepath) else 0
        if file_size > CONFIG.max_file_size_bytes:
            logging.warning(f"{os.path.basename(filepath)} too large ({file_size} bytes), skipping.")
            return []
        if file_size > CONFIG.LARGE_FILE_THRESHOLD_BYTES:
            matches, all_contexts = process_large_file_chunked(filepath, keywords, regex_mode, exact_match, session_id, privacy_logger)
            file_hash = calculate_file_hash(filepath)
            stat = os.stat(filepath)
            file_info = {
                'size': stat.st_size,
                'modified': datetime.datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                'created': datetime.datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                'hash': file_hash,
                'pii_types_detected': [], # Limited for chunked
                'pii_masked': bool(privacy_logger and PRIVACY_SETTINGS["enable_pii_masking"])
            }
            anonymized_filename = anonymize_file_path(display_name)
            return [ [anonymized_filename, matches, all_contexts, file_info, {}] ]

    # v5.0: Buffer stream before reading to ensure seekable hash calculation (BUG-21)
    if file_obj is not None:
        try:
            raw_bytes = file_obj.read()
            file_hash = hashlib.sha256(raw_bytes).hexdigest()
            file_obj = io.BytesIO(raw_bytes)   # seekable buffer for read_file_content
        except Exception as e:
            logging.warning(f"Stream tampolanamadı ({original_name or filepath}): {e}")
            file_hash = "Hash-Error"
    else:
        file_hash = None  # will be calculated below for physical files

    content = read_file_content(filepath, file_stream=file_obj, file_ext=ext)
    
    if not content:
        return []
    
    detected_pii = detect_pii(content)
    if detected_pii and privacy_logger:
        log_privacy_event(privacy_logger, "PII_DETECTED", f"File: {display_name}, Types: {list(detected_pii.keys())}", session_id)
    
    if PRIVACY_SETTINGS["enable_pii_masking"] and detected_pii and privacy_logger:
        content = mask_pii(content, detected_pii, session_id, privacy_logger)
    
    matches = {}
    all_contexts = []
    
    # v5.0: Hash already calculated above for streams (BUG-21 fix)
    if file_hash is None:
        file_hash = calculate_file_hash(filepath)
    
    if file_obj is not None:
        file_info = {
            'size': member_size,
            'modified': 'N/A (Archive)',
            'created': 'N/A (Archive)',
            'hash': file_hash,
            'pii_types_detected': list(detected_pii.keys()) if detected_pii else [],
            'pii_masked': bool(detected_pii and PRIVACY_SETTINGS["enable_pii_masking"])
        }
    else:
        stat = os.stat(filepath)
        file_info = {
            'size': stat.st_size,
            'modified': datetime.datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
            'created': datetime.datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
            'hash': file_hash,
            'pii_types_detected': list(detected_pii.keys()) if detected_pii else [],
            'pii_masked': bool(detected_pii and PRIVACY_SETTINGS["enable_pii_masking"])
        }
    
    keyword_patterns = []
    for keyword in keywords:
        kw = keyword.strip()
        try:
            if regex_mode:
                pattern = re.compile(kw, re.IGNORECASE)
            elif exact_match:
                pattern = re.compile(r'\b' + re.escape(kw) + r'\b', re.IGNORECASE)
            else:
                pattern = re.compile(re.escape(kw), re.IGNORECASE)
            keyword_patterns.append((kw, pattern))
        except re.error: continue

    for kw_str, pattern in keyword_patterns:
        # v5.0: ReDoS protected search
        matches_found = safe_regex_search(pattern, content)
        count = len(matches_found)
        
        if count > 0:
            matches[kw_str] = count
            contexts = extract_context(content, pattern, apply_masking=True, session_id=session_id, privacy_logger=privacy_logger, highlight=True)
            all_contexts.extend(contexts)
    
    anonymized_filename = anonymize_file_path(display_name)
    return [ [anonymized_filename, matches, all_contexts, file_info, detected_pii] ]

# The extract_context_v52 was a temporary name, now merged back.
# (Removing the duplicate function that was added in previous step)
def search_keywords(
    directory: str, 
    keywords: List[str], 
    exact_match: bool = False, 
    regex_mode: bool = False, 
    recursive: bool = False, 
    session_id: str = "", 
    audit_logger: Optional[logging.Logger] = None, 
    privacy_logger: Optional[logging.Logger] = None
) -> Tuple[Dict, int, Dict, List[str]]:
    """
    TraceWords: Parallel scan files in a folder and return digital evidence.
    v5.0: Threading optimized with as_completed.
    """
    from concurrent.futures import as_completed
    found = {}
    pii_summary = {}
    start_time = time.time()
    
    # Reset archive stats for new session
    archive_stats.reset()

    logging.info(f"TraceWords analizi başlatıldı: {directory}")
    CONSOLE.print(f"[bold blue]🔍 TraceWords dijital anahtar kelime analizi:[/bold blue] [yellow]{directory}[/yellow]")
    
    if audit_logger:
        log_audit_event(audit_logger, "ANALYSIS_STARTED", directory, "user", session_id)
    
    if not os.path.exists(directory):
        logging.error("Hedef dizin bulunamadı!")
        CONSOLE.print("[bold red]❌ Hedef dizin bulunamadı![/bold red]")
        if audit_logger:
            log_audit_event(audit_logger, "ANALYSIS_FAILED", directory, "user", session_id, "DIRECTORY_NOT_FOUND")
        return found, 0, pii_summary, []
    
    # Create file list
    files = []
    if os.path.isfile(directory):
        if directory.lower().endswith(CONFIG.SUPPORTED_FORMATS):
            files.append(directory)
    elif os.path.isdir(directory):
        if recursive:
            for root, dirs, filenames in os.walk(directory):
                for filename in filenames:
                    if filename.lower().endswith(CONFIG.SUPPORTED_FORMATS):
                        files.append(os.path.join(root, filename))
        else:
            files = [
                os.path.join(directory, f) for f in os.listdir(directory)
                if f.lower().endswith(CONFIG.SUPPORTED_FORMATS)
            ]
    else:
        CONSOLE.print(f"[bold red]❌ Geçersiz yol (Dosya veya dizin değil): {directory}[/bold red]")
        return found, 0, pii_summary, []
    
    file_count = len(files)
    
    if not files:
        CONSOLE.print(f"[yellow]⚠️  Desteklenen formatlarda dosya bulunamadı: {', '.join(CONFIG.SUPPORTED_FORMATS)}[/yellow]")
        logging.warning("Analiz edilecek dosya bulunamadı.")
        if audit_logger:
            log_audit_event(audit_logger, "ANALYSIS_COMPLETED", directory, "user", session_id, "NO_FILES_FOUND")
        return found, 0, pii_summary, []
    
    CONSOLE.print(f"[cyan]📁 Desteklenen formatlar:[/cyan] [white]{', '.join(CONFIG.SUPPORTED_FORMATS)}[/white]")
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        TextColumn("[blue]({task.completed}/{task.total})"),
        console=CONSOLE
    ) as progress:
        scan_task = progress.add_task("[green]Dijital kanıt analizi...", total=file_count)
        
        # v5.0 Dynamic workers: Optimized for both I/O and CPU
        cpu_count = multiprocessing.cpu_count()
        io_workers = min(32, cpu_count * 4)
        
        with ThreadPoolExecutor(max_workers=io_workers) as executor:
            futures = {
                executor.submit(process_file, filepath, keywords, exact_match, regex_mode, session_id, privacy_logger): filepath
                for filepath in files
            }
            
            for future in as_completed(futures):
                file_name = futures[future]
                try:
                    # v5.0: Added timeout for each task (BUG-09-B)
                    results_list = future.result(timeout=CONFIG.FILE_SCAN_TIMEOUT_SEC)
                    
                    # Results are ALWAYs a list of [filepath, matches, contexts, file_info, detected_pii]
                    for res in results_list:
                        res_filepath, matches, contexts, file_info, detected_pii = res
                        if matches:
                            found[res_filepath] = {
                                "matches": matches, 
                                "contexts": contexts,
                                "file_info": file_info
                            }
                        
                        # Create PII summary
                        if detected_pii:
                            for pii_type, pii_values in detected_pii.items():
                                if pii_type not in pii_summary:
                                    pii_summary[pii_type] = 0
                                pii_summary[pii_type] += len(pii_values)
                except TimeoutError:
                    logging.warning(f"Zaman aşımı: {file_name} {CONFIG.FILE_SCAN_TIMEOUT_SEC}s içinde tamamlanamadı, atlanıyor.")
                    CONSOLE.print(f"[yellow]⏱ Zaman aşımı:[/yellow] {os.path.basename(file_name)} atlandı.")
                except Exception as e:
                    logging.error(f"Hata: {file_name} işlenirken hata oluştu: {e}")
                
                progress.update(scan_task, advance=1)
    
    analysis_time = time.time() - start_time
    
    table = Table(title="Analiz İstatistikleri", show_header=True, header_style="bold magenta")
    table.add_column("Metrik", style="cyan")
    table.add_column("Değer", style="green")
    
    table.add_row("Analiz Edilen Dosya", str(file_count))
    table.add_row("Analiz Süresi", f"{analysis_time:.2f} saniye")
    table.add_row("Kanıt Bulunan Dosya", str(len(found)))
    
    if pii_summary:
        table.add_row("Tespit Edilen PII Türleri", ", ".join(pii_summary.keys()))
        if privacy_logger:
            log_privacy_event(privacy_logger, "PII_SUMMARY", f"Types: {pii_summary}", session_id)
    
    CONSOLE.print(table)
    
    if found:
        CONSOLE.print("\n[bold green]✅ Önemli Bulgular:[/bold green]")
        finding_table = Table(show_header=True, header_style="bold yellow")
        finding_table.add_column("Dosya (Anonim/Orijinal)", style="dim")
        finding_table.add_column("Eşleşmeler", style="bold")
        finding_table.add_column("PII", style="red")
        
        for file, data in list(found.items())[:10]:
            matches_str = ", ".join([f"{k} ({v})" for k, v in data["matches"].items()])
            pii_str = "⚠️" if data["file_info"]["pii_types_detected"] else "✅"
            finding_table.add_row(os.path.basename(file), matches_str, pii_str)
        
        CONSOLE.print(finding_table)
        
        # Önizleme (Preview)
        CONSOLE.print("\n[bold cyan]🔍 Kanıt Önizlemesi (En Çok Eşleşen Dosya):[/bold cyan]")
        top_file = sorted(found.items(), key=lambda x: sum(x[1]["matches"].values()), reverse=True)[0]
        filepath, data = top_file
        CONSOLE.print(f"[yellow]📄 {os.path.basename(filepath)}[/yellow]")
        for i, ctx in enumerate(data["contexts"][:3], 1):
            CONSOLE.print(f"  [dim]Satır {ctx['line_number']}:[/dim] {ctx['matched_line']}")
        
        if len(found) > 10:
            CONSOLE.print(f"\n[dim]...ve {len(found)-10} dosya daha bulundu. Tüm detaylar rapor dosyasında.[/dim]")
    
    logging.info(f"TraceWords analizi tamamlandı. Dosya sayısı: {file_count}, Süre: {analysis_time:.2f} saniye")
    
    if audit_logger:
        log_audit_event(audit_logger, "ANALYSIS_COMPLETED", f"{directory} - {file_count} files", "user", session_id)
    
    return found, file_count, pii_summary, files

def strip_rich_tags(text: str) -> str:
    """Clean rich markup tags"""
    return re.sub(r'\[/?[a-z ]+\]', '', text)

def save_report(
    results: Dict, 
    output_file: str, 
    directory: str, 
    keywords: List[str], 
    session_id: str, 
    pii_summary: Dict, 
    audit_logger: Optional[logging.Logger] = None, 
    privacy_logger: Optional[logging.Logger] = None, 
    batch_mode: bool = False, 
    encrypt: bool = False
) -> None:
    """Save TraceWords analysis results with path traversal protection and optional encryption"""
    # Path Traversal Protection: Ensure report stays in the intended directory
    output_file = os.path.basename(output_file)
    target_dir = os.path.dirname(directory) if os.path.isfile(directory) else directory
    full_path = os.path.join(target_dir, output_file)
    
    if os.path.exists(full_path) and not batch_mode:
        overwrite = questionary.confirm(f"{output_file} zaten mevcut. Üzerine yazılsın mı?").ask()
        if not overwrite:
            CONSOLE.print("[yellow]📄 Rapor kaydedilmedi.[/yellow]")
            logging.info(f"Rapor kaydedilmedi: {full_path}")
            if audit_logger:
                log_audit_event(audit_logger, "REPORT_SAVE_CANCELLED", full_path, "user", session_id)
            return
    
    report_content = []
    try:
        # Generate report content
        report_content.append("=" * 80 + "\n")
        report_content.append("TRACEWORDS — GDPR/CCPA UYUMLU DİJİTAL ADLİ TARAMA RAPORU\n")
        report_content.append("=" * 80 + "\n")
        report_content.append("TraceWords v5.0 - GDPR/CCPA Uyumlu Dijital Adli Tarama Aracı\n")
        report_content.append(f"Analiz Tarihi: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        report_content.append(f"Oturum ID: {session_id}\n")
        report_content.append(f"Aranan Terimler: {', '.join(keywords)}\n")
        report_content.append(f"Bulunan Dijital Kanıt Dosyası: {len(results)}\n")
        report_content.append(f"Analiz Edilen Dizin: {directory}\n\n")
        
        report_content.append("VERİ GİZLİLİĞİ VE UYUMLULUK BİLGİLERİ\n")
        report_content.append("-" * 50 + "\n")
        report_content.append(f"PII Maskeleme Durumu: {'Aktif' if PRIVACY_SETTINGS['enable_pii_masking'] else 'Pasif'}\n")
        report_content.append(f"Veri Minimizasyonu: {'Aktif' if PRIVACY_SETTINGS['enable_data_minimization'] else 'Pasif'}\n")
        report_content.append(f"Audit Loglama: {'Aktif' if PRIVACY_SETTINGS['enable_audit_logging'] else 'Pasif'}\n")
        report_content.append(f"Sonuç Anonimleştirme: {'Aktif' if PRIVACY_SETTINGS['anonymize_results'] else 'Pasif'}\n")
        report_content.append(f"Veri Saklama Süresi: {PRIVACY_SETTINGS['retention_period_days']} gün\n")
        report_content.append(f"Silinme Tarihi: {(datetime.datetime.now() + datetime.timedelta(days=PRIVACY_SETTINGS['retention_period_days'])).strftime('%Y-%m-%d')}\n\n")
        
        if pii_summary:
            report_content.append("KİŞİSEL VERİ (PII) TESPİT ÖZETİ\n")
            report_content.append("-" * 40 + "\n")
            for pii_type, count in pii_summary.items():
                pii_type_tr = {
                    'email': 'E-posta Adresi', 'phone_us': 'ABD Telefon No', 'phone_international': 'Uluslararası Telefon',
                    'ssn_us': 'ABD SSN', 'credit_card': 'Kredi Kartı No', 'ip_address': 'IP Adresi',
                    'mac_address': 'MAC Adresi', 'tc_kimlik': 'T.C. Kimlik No', 'iban': 'IBAN No',
                    'passport': 'Pasaport No', 'date_of_birth': 'Doğum Tarihi'
                }.get(pii_type, pii_type.upper())
                report_content.append(f"  🔒 {pii_type_tr}: {count} adet\n")
            report_content.append("\n⚠️  Tüm kişisel veriler maskelenmiştir ve orijinal değerler rapordan çıkarılmıştır.\n\n")
        
        if results:
            total_matches = 0
            for file, data in results.items():
                matches = data["matches"]; contexts = data["contexts"]; file_info = data["file_info"]
                report_content.append("-" * 70 + "\n")
                report_content.append(f"DİJİTAL KANIT DOSYASI: {file}\n")
                report_content.append("-" * 70 + "\n")
                report_content.append(f"Dosya Boyutu: {file_info['size']:,} bytes\n")
                report_content.append(f"Son Değiştirilme: {file_info['modified']}\n")
                report_content.append(f"Oluşturulma Tarihi: {file_info['created']}\n")
                report_content.append(f"SHA-256 Hash (Adli Bütünlük): {file_info['hash']}\n")
                
                if file_info.get('pii_types_detected'):
                    report_content.append(f"Tespit Edilen PII Türleri: {', '.join(file_info['pii_types_detected'])}\n")
                
                report_content.append("\nBULUNAN DİJİTAL KANITLAR:\n")
                for keyword, count in matches.items():
                    report_content.append(f"  🔍 {keyword}: {count} eşleşme\n")
                    total_matches += count
                
                if contexts:
                    report_content.append("\nKANIT BAĞLAMLARI (PII MASKELENMİŞ):\n")
                    for i, context in enumerate(contexts[:5], 1):
                        report_content.append(f"\n  [{i}] Satır {context['line_number']}:\n")
                        report_content.append(f"      Eşleşen: {strip_rich_tags(context['matched_line'])}\n")
                        report_content.append(f"  Bağlam:\n")
                        for line in context['context'].split('\n'):
                            report_content.append(f"      {strip_rich_tags(line)}\n")
                report_content.append("\n" + "=" * 70 + "\n\n")
            
            report_content.append("TRACEWORDS ANALİZ ÖZETİ\n")
            report_content.append("=" * 30 + "\n")
            report_content.append(f"Toplam Dijital Kanıt: {total_matches}\n")
            report_content.append(f"Kanıt İçeren Dosya Sayısı: {len(results)}\n")
        else:
            report_content.append("❌ HİÇBİR DİJİTAL KANIT BULUNAMADI.\n")
        
        report_content.append("\nGDPR/CCPA UYUMLULUK BİLDİRİMİ\n" + "=" * 30 + "\n")
        report_content.append("Bu rapor TraceWords v5.0 tarafından veri gizliliği standartlarına uygun hazırlanmıştır.\n")
        
        final_text = "".join(report_content)
        if encrypt:
            storage = GDPRCompliantStorage()
            if storage.store_encrypted(final_text, full_path + ".enc"):
                CONSOLE.print(f"[green]🔒 Şifreli rapor kaydedildi: {output_file}.enc[/green]")
                if audit_logger:
                    log_audit_event(audit_logger, "REPORT_SAVED", full_path + ".enc", "user", session_id)
            else:
                CONSOLE.print(f"[red]HATA: Şifreli rapor kaydedilemedi![/red]")
                logging.error(f"Şifreli rapor kaydedilemedi: {full_path}")
                if audit_logger:
                    log_audit_event(audit_logger, "REPORT_SAVE_FAILED", full_path + ".enc", "user", session_id, "ERROR")
        else:
            with open(full_path, "w", encoding="utf-8") as report:
                report.write(final_text)
            CONSOLE.print(f"[green]📄 Rapor kaydedildi: {output_file}[/green]")
            if audit_logger:
                log_audit_event(audit_logger, "REPORT_SAVED", full_path, "user", session_id)
            
    except Exception as e:
        logging.error(f"Rapor kaydedilemedi: {e}", exc_info=True)
        CONSOLE.print(f"[bold red]❌ Rapor kaydedilemedi: {e}[/bold red]")

def cleanup_old_data():
    """Clean up old log files using secure delete"""
    retention_days = PRIVACY_SETTINGS["retention_period_days"]
    cutoff_date = datetime.datetime.now() - datetime.timedelta(days=retention_days)
    cleanup_count = 0
    log_files = [CONFIG.LOG_FILE, CONFIG.AUDIT_LOG_FILE, CONFIG.PRIVACY_LOG_FILE]
    
    for log_file in log_files:
        if os.path.exists(log_file):
            file_time = datetime.datetime.fromtimestamp(os.path.getmtime(log_file))
            if file_time < cutoff_date:
                try:
                    secure_delete(log_file)
                    cleanup_count += 1
                    CONSOLE.print(f"[yellow]🗑️  Eski veri dosyası güvenli silindi: {log_file}[/yellow]")
                except Exception as e:
                    logging.error(f"Hata: {log_file} silinemedi: {e}")
    if cleanup_count > 0:
        CONSOLE.print(f"[green]✅ GDPR/CCPA uyumlu veri temizliği: {cleanup_count} dosya silindi[/green]")
    return cleanup_count

def parse_args():
    """Parse command line arguments - v5.0: CLI Flag Improvements."""
    parser = argparse.ArgumentParser(
        description="TraceWords v5.0 — GDPR/CCPA Uyumlu Gelişmiş Dijital Adli Tarama Aracı",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument(
        "directory",
        help="Analiz edilecek dizin yolu"
    )
    parser.add_argument(
        "-k", "--keywords",
        help="Virgülle ayrılmış arama terimleri (Opsiyonel, etkileşimli girilebilir)"
    )
    
    # Mutually Exclusive Search Modes
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument(
        "-e", "--exact",
        action="store_true",
        help="Tam kelime eşleşmesi (varsayılan: False)"
    )
    mode_group.add_argument(
        "-r", "--regex",
        action="store_true",
        help="Regex pattern arama modu"
    )
    
    parser.add_argument(
        "--recursive",
        action="store_true",
        help="Alt dizinleri de analiz et"
    )
    parser.add_argument(
        "-o", "--output",
        default="tracewords_report.txt",
        help="Rapor dosya adı (varsayılan: tracewords_report.txt)"
    )
    parser.add_argument(
        "--no-pii-mask",
        action="store_true",
        help="PII maskelemeyi devre dışı bırak"
    )
    parser.add_argument(
        "--anonymize",
        action="store_true",
        help="Dosya isimlerini anonimleştir"
    )
    parser.add_argument(
        "--cleanup",
        action="store_true",
        help="Eski verileri temizle (loglar ve raporlar)"
    )
    parser.add_argument(
        "--encrypt",
        action="store_true",
        help="Raporu şifreli olarak kaydet"
    )
    parser.add_argument(
        "--self-test",
        action="store_true",
        help="Sistem güvenlik ve doğrulama testlerini çalıştırır"
    )
    parser.add_argument(
        "-b", "--batch",
        action="store_true",
        help="Otomasyon (batch) modu: Etkileşimli soruları atlar"
    )
    parser.add_argument(
        "--encrypt-logs",
        action="store_true",
        help="Log dosyalarını da şifreli olarak saklar"
    )
    parser.add_argument(
        "--wipe-source",
        action="store_true",
        help="Analiz sonrası kaynak dosyaları güvenli bir şekilde siler (DİKKAT!)"
    )
    return parser.parse_args()

def run_self_tests():
    """Run built-in security and sanity tests for v5.0 features"""
    CONSOLE.print("[bold blue]🧪 TraceWords v5.0 Self-Test Başlatılıyor...[/bold blue]\n")
    
    # 1. Path Traversal
    CONSOLE.print("1. Path Traversal Koruması Test Ediliyor...")
    test_paths = [("../../etc/passwd", "passwd"), ("C:/Windows/System32/config/SAM", "SAM")]
    for p, expected in test_paths:
        res = sanitize_path(p)
        if res == expected:
            CONSOLE.print(f"  [green]✅ PASSED:[/green] {p} -> {res}")
        else:
            CONSOLE.print(f"  [red]❌ FAILED:[/red] {p} -> {res} (Beklenen: {expected})")

    # 2. TC Kimlik (BUG-06 Optimized)
    CONSOLE.print("\n2. TC Kimlik Doğrulama Algoritması Test Ediliyor...")
    tcs = [("10000000146", True), ("10000000147", False), ("1000000014", False), ("A10000000146Z", False)]
    for tc, expected in tcs:
        res = validate_tc_kimlik(tc)
        if res == expected:
            CONSOLE.print(f"  [green]✅ PASSED:[/green] {tc} -> {res}")
        else:
            CONSOLE.print(f"  [red]❌ FAILED:[/red] {tc} -> {res}")

    # 3. Keyword Validation
    CONSOLE.print("\n3. Anahtar Kelime Doğrulama Test Ediliyor...")
    kw_tests = [
        (["test", "valid"], True),
        (["a"], False),
        (["test; rm -rf"], False),
        ([], False)
    ]
    for kws, expected in kw_tests:
        res, msg = validate_keywords(kws)
        if res == expected:
            CONSOLE.print(f"  [green]✅ PASSED:[/green] {kws} -> {res} {msg}")
        else:
            CONSOLE.print(f"  [red]❌ FAILED:[/red] {kws} -> {res} {msg}")

    # 4. Encryption Roundtrip (BUG-11-A)
    CONSOLE.print("\n4. GDPR Şifreleme Roundtrip Test Ediliyor...")
    try:
        storage = GDPRCompliantStorage()
        test_data = "Secret Data — Gizli Test Verisi"
        test_file = "test_encrypted.bin"
        
        storage.store_encrypted(test_data, test_file)
        
        if os.path.exists(test_file):
            # Verify roundtrip
            with open(test_file, 'rb') as f:
                encrypted_bytes = f.read()
            decrypted = storage.decrypt_data(encrypted_bytes)
            
            os.remove(test_file)
            
            if decrypted == test_data:
                CONSOLE.print("  [green]✅ PASSED:[/green] Şifreleme + çözme roundtrip başarılı.")
            else:
                CONSOLE.print(f"  [red]❌ FAILED:[/red] Roundtrip başarısız (Beklenen: {test_data}, Alınan: {decrypted})")
        else:
            CONSOLE.print("  [red]❌ FAILED:[/red] Dosya oluşturulamadı.")
    except Exception as e:
        CONSOLE.print(f"  [red]❌ FAILED:[/red] Şifreleme hatası: {e}")
        if os.path.exists(test_file): os.remove(test_file)

    CONSOLE.print("\n[bold green]✅ Tüm self-testler tamamlandı.[/bold green]")

def main():
    """Main program flow - v5.0: Hardened orchestration."""
    # Centralized logging initialization happens only once here if called directly
    audit_logger, privacy_logger = setup_logging()
    logging.info("TraceWords v5.0 programı başlatıldı.")
    
    session_id = generate_session_id()
    args = parse_args()
    
    # Batch mode support for cleanup
    if args.cleanup:
        cleanup_old_data()
        return

    if args.self_test:
        run_self_tests()
        return
    
    # Update Privacy settings
    if args.no_pii_mask:
        PRIVACY_SETTINGS["enable_pii_masking"] = False
    if args.anonymize:
        PRIVACY_SETTINGS["anonymize_results"] = True
    
    # GDPR/CCPA Consent check
    if not check_consent(session_id, audit_logger, batch_mode=args.batch):
        return
    
    if not os.path.exists(args.directory):
        CONSOLE.print("❌ Hedef dizin bulunamadı!")
        logging.error(f"Hedef dizin bulunamadı: {args.directory}")
        log_audit_event(audit_logger, "DIRECTORY_NOT_FOUND", args.directory, "user", session_id, "ERROR")
        return
    
    # Keyword acquisition: Support both CLI and interactive
    keywords = []
    if args.keywords:
        # v5.0 BUG FIX: Regex mode logic for commas (BUG-18)
        if args.regex and "," in args.keywords and (args.keywords.count("[") > 0 or args.keywords.count("{") > 0):
            # If it looks like a complex regex, don't split by comma
            keywords = [args.keywords.strip()]
        else:
            keywords = [kw.strip() for kw in args.keywords.split(",") if kw.strip()]
    
    if not keywords and not args.batch:
        kw_input = questionary.text("Analiz edilecek anahtar kelimeleri girin (virgülle ayırın):").ask()
        if kw_input:
            keywords = [kw.strip() for kw in kw_input.split(",") if kw.strip()]
            
    # Validate keywords
    valid, error_msg = validate_keywords(keywords)
    if not valid:
        CONSOLE.print(f"[bold red]❌ Geçersiz Anahtar Kelime(ler):[/bold red] {error_msg}")
        logging.error(f"Invalid keywords: {error_msg}")
        log_audit_event(audit_logger, "INVALID_KEYWORDS", error_msg, "user", session_id, "ERROR")
        return
    
    # Determine search mode
    exact_match = args.exact
    regex_mode = args.regex
    recursive = args.recursive
    
    # Interactive mode fallback for search mode
    if not args.exact and not args.regex and not args.batch:
        choice = questionary.select(
            "🔍 TraceWords arama modu seçin:",
            choices=[
                "1. Kısmi eşleşme (varsayılan)",
                "2. Tam kelime eşleşmesi",
                "3. Regex pattern arama"
            ]
        ).ask()
        
        if choice and choice.startswith("2"):
            exact_match = True
        elif choice and choice.startswith("3"):
            regex_mode = True
    
    param_panel = Panel(
        f"[cyan]📂 Hedef Dizin:[/cyan] {args.directory}\n"
        f"[cyan]🔍 Arama Terimleri:[/cyan] {', '.join(keywords)}\n"
        f"[cyan]✅ Tam Eşleşme:[/cyan] [yellow]{'Evet' if exact_match else 'Hayır'}[/yellow]\n"
        f"[cyan]🔄 Regex Modu:[/cyan] [yellow]{'Evet' if regex_mode else 'Hayır'}[/yellow]\n"
        f"[cyan]📁 Recursive:[/cyan] [yellow]{'Evet' if recursive else 'Hayır'}[/yellow]\n"
        f"[cyan]🔒 PII Maskeleme:[/cyan] [yellow]{'Evet' if PRIVACY_SETTINGS['enable_pii_masking'] else 'Hayır'}[/yellow]\n"
        f"[cyan]🎭 Anonimleştirme:[/cyan] [yellow]{'Evet' if PRIVACY_SETTINGS['anonymize_results'] else 'Hayır'}[/yellow]\n"
        f"[cyan]📋 Desteklenen Formatlar:[/cyan] {len(CONFIG.SUPPORTED_FORMATS)} adet\n"
        f"[cyan]🆔 Oturum ID:[/cyan] [dim]{session_id}[/dim]",
        title="[bold blue]🔧 Analiz Parametreleri[/bold blue]", border_style="blue"
    )
    CONSOLE.print(param_panel)
    
    log_audit_event(audit_logger, "ANALYSIS_PARAMETERS", f"Keywords: {keywords}, Mode: {'regex' if regex_mode else 'exact' if exact_match else 'partial'}", "user", session_id)
    
    CONSOLE.print("\n[bold green]🚀 TraceWords GDPR/CCPA uyumlu dijital analiz başlatılıyor...[/bold green]\n")
    results, file_count, pii_summary, files = search_keywords(
        args.directory, keywords, exact_match, regex_mode, recursive, 
        session_id, audit_logger, privacy_logger
    )
    
    if results:
        CONSOLE.print(f"\n✅ {len(results)} dosyada dijital kanıt bulundu!")
        save_report(results, args.output, args.directory, keywords, session_id, pii_summary, audit_logger, privacy_logger, batch_mode=args.batch, encrypt=args.encrypt)
    else:
        CONSOLE.print("\n[bold yellow]❌ Hiçbir dijital kanıt bulunamadı.[/bold yellow]")
        log_audit_event(audit_logger, "NO_EVIDENCE_FOUND", args.directory, "user", session_id)

    # DÜZELTME: wipe_source if results'ın dışında, tüm taranan dosyalar için (BUG-08-B)
    if args.wipe_source and files:
        confirm_wipe = args.batch or questionary.confirm(f"DİKKAT: {len(files)} kaynak dosya GÜVENLİ olarak silinecek. Emin misiniz?").ask()
        if confirm_wipe:
            with Progress(console=CONSOLE) as progress:
                wipe_task = progress.add_task("[red]Kaynak dosyalar güvenli siliniyor...", total=len(files))
                for filepath in files:
                    secure_delete(filepath)
                    progress.update(wipe_task, advance=1)
            CONSOLE.print("[bold red]🗑️  Tüm kaynak dosyalar güvenli silindi.[/bold red]")
            log_audit_event(audit_logger, "SOURCE_WIPE_COMPLETED", args.directory, "user", session_id)
    elif args.wipe_source and not files:
        CONSOLE.print("[yellow]⚠ --wipe-source: Silinecek dosya bulunamadı.[/yellow]")

    # Log encryption logic (at the very end)
    if args.encrypt_logs:
        CONSOLE.print("[cyan]🔒 Log dosyaları şifreleniyor...[/cyan]")
        storage = GDPRCompliantStorage()
        all_success = True
        for log_file in [CONFIG.LOG_FILE, CONFIG.AUDIT_LOG_FILE, CONFIG.PRIVACY_LOG_FILE]:
            if os.path.exists(log_file):
                with open(log_file, 'rb') as lf:
                    data = lf.read()
                enc_path = log_file + ".enc"
                success = storage.store_encrypted_bytes(data, enc_path)
                if success and os.path.exists(enc_path):
                    secure_delete(log_file)
                else:
                    all_success = False
                    logging.error(f"Log şifrelenemedi, orijinal korunuyor: {log_file}")
                    CONSOLE.print(f"[red]Uyarı: {log_file} şifrelenemedi.[/red]")
        if all_success:
            CONSOLE.print("[green]✅ Loglar şifrelendi ve orijinalleri güvenli silindi.[/green]")
        else:
            CONSOLE.print("[yellow]⚠ Bazı log dosyaları şifrelenemedi. Detaylar için log dosyasını inceleyin.[/yellow]")

if __name__ == "__main__":
    welcome_panel = Panel(
        "[bold cyan]TRACEWORDS v5.0[/bold cyan]\n"
        "[yellow]GDPR/CCPA Uyumlu Gelişmiş Dijital Adli Tarama Aracı[/yellow]\n\n"
        "[green]🔒 Veri Gizliliği Yasalarına (GDPR/CCPA) Tam Uyum[/green]\n"
        "[white]• Endüstriyel Ölçekte Paralel Analiz Motoru[/white]\n"
        "[white]• RAM Tasarruflu Archive Streaming Desteği[/white]",
        border_style="cyan"
    )
    CONSOLE.print(welcome_panel)
    CONSOLE.print(f"[dim]🕐 Başlangıç zamanı: {datetime.datetime.now().strftime('%H:%M:%S')}[/dim]")
    
    try:
        main()
    except KeyboardInterrupt:
        CONSOLE.print("\n\n[bold yellow]⚠️  TraceWords analizi kullanıcı tarafından iptal edildi.[/bold yellow]")
        logging.warning("TraceWords analizi kullanıcı tarafından iptal edildi.")
    except Exception as e:
        CONSOLE.print(f"\n[bold red]❌ Beklenmeyen hata: {e}[/bold red]")
        logging.error(f"Beklenmeyen hata: {e}", exc_info=True)
    finally:
        CONSOLE.print(f"\n[bold blue]🏁 TraceWords analizi sona erdi[/bold blue] - [dim]{datetime.datetime.now().strftime('%H:%M:%S')}[/dim]")
        CONSOLE.print("[cyan]📋 Loglar:[/cyan] [white]tracewords_info.log, tracewords_audit.log, tracewords_privacy.log[/white]")
        logging.info("TraceWords v5.0 programı sona erdi.")

